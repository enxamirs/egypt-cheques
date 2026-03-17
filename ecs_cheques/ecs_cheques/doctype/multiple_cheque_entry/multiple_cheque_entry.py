# Copyright (c) 2021, erpcloud.systems and contributors
# For license information, please see license.txt

import frappe
import io
from frappe.model.document import Document
from frappe import _
from frappe.utils import flt, nowdate, getdate


# ---------------------------------------------------------------------------
# Exchange-rate / amount helpers
# ---------------------------------------------------------------------------

def _get_account_currency_db(account_name, company_currency):
	"""Return the account's currency from the Account master, falling back to
	*company_currency* if the account is blank or not found."""
	if not account_name:
		return company_currency
	return frappe.db.get_value("Account", account_name, "account_currency") or company_currency


def _fetch_exchange_rate_to_company(from_currency, company_currency, posting_date=None):
	"""Fetch the exchange rate from *from_currency* to *company_currency* using the
	Currency Exchange DocType.

	Returns 1.0 when currencies are equal.  Searches for the most recent rate on or
	before *posting_date*.  Tries the direct pair first, then the inverse pair.
	Returns None when no matching record is found.
	"""
	if not from_currency or from_currency == company_currency:
		return 1.0

	date_filter = getdate(posting_date) if posting_date else getdate(nowdate())

	# Direct: from_currency → company_currency
	rate = frappe.db.get_value(
		"Currency Exchange",
		{"from_currency": from_currency, "to_currency": company_currency,
		 "date": ["<=", date_filter]},
		"exchange_rate",
		order_by="date desc",
	)
	if rate:
		return flt(rate)

	# Inverse: company_currency → from_currency
	rate = frappe.db.get_value(
		"Currency Exchange",
		{"from_currency": company_currency, "to_currency": from_currency,
		 "date": ["<=", date_filter]},
		"exchange_rate",
		order_by="date desc",
	)
	if rate and flt(rate) > 0:
		return flt(1.0 / flt(rate), 9)

	return None


@frappe.whitelist()
def get_exchange_rate_to_company(from_currency, company, posting_date=None):
	"""Whitelisted API: return the exchange rate from *from_currency* to the
	company's default currency.  Used by the client-side ``update_amount_in_usd``
	helper to avoid having to inline the Currency Exchange lookup in JS.

	Returns the numeric rate (or 1.0 as a fallback) so callers can always
	multiply: ``amount_in_company_currency = paid_amount × rate``.
	"""
	company_currency = (
		frappe.db.get_value("Company", company, "default_currency") or ""
		if company else
		frappe.db.get_default("currency") or ""
	)
	if not company_currency or from_currency == company_currency:
		return 1.0
	rate = _fetch_exchange_rate_to_company(from_currency, company_currency, posting_date)
	return rate if rate is not None else 1.0


def _compute_payment_entry_amounts(
	row_paid_amount,
	paid_from_currency,
	paid_to_currency,
	company_currency,
	stored_exchange_rate,
	payment_type,
):
	"""Compute ``paid_amount``, ``received_amount``, ``source_exchange_rate`` and
	``target_exchange_rate`` for a Payment Entry from **actual** account currencies.

	ERPNext v15 conventions
	-----------------------
	* ``paid_amount``         – amount in ``paid_from_account_currency``
	* ``received_amount``     – amount in ``paid_to_account_currency``
	* ``source_exchange_rate`` – ``paid_from_currency`` → ``company_currency``
	* ``target_exchange_rate`` – ``paid_to_currency``   → ``company_currency``

	``row_paid_amount`` is in:

	* **Receive** → ``paid_to_currency``   (cheque/bank amount, e.g. 1 000 USD)
	* **Pay**     → ``paid_from_currency`` (cheque/bank amount, e.g. 1 000 USD)

	``stored_exchange_rate`` is the child-table ``target_exchange_rate`` value,
	which represents the foreign-currency → company-currency conversion rate
	(e.g. 3.159059 for USD → ILS when company currency is ILS).
	"""
	rate = flt(stored_exchange_rate) or 1.0
	raw_amount = flt(row_paid_amount)

	if paid_from_currency == paid_to_currency:
		return dict(
			paid_from_account_currency=paid_from_currency,
			paid_to_account_currency=paid_to_currency,
			paid_amount=raw_amount,
			received_amount=raw_amount,
			source_exchange_rate=1.0,
			target_exchange_rate=1.0,
		)

	# Determine each side's exchange rate to company currency.
	# For Receive: stored_exchange_rate = paid_to_currency → paid_from_currency (bank → party)
	# For Pay:     stored_exchange_rate = paid_from_currency → paid_to_currency  (bank → party)
	# When company_currency matches paid_from or paid_to, derive the other rate accordingly.
	if paid_from_currency == company_currency:
		source_exchange_rate = 1.0
		if payment_type == "Receive":
			# stored = paid_to (foreign) → paid_from (company) = foreign → company ✓
			target_exchange_rate_pe = rate
		else:
			# Pay: paid_from = bank = company; stored = bank (company) → paid_to (foreign).
			# target = paid_to → company = 1 / stored
			target_exchange_rate_pe = flt(1.0 / rate, 9) if rate else 1.0
	elif paid_to_currency == company_currency:
		target_exchange_rate_pe = 1.0
		if payment_type == "Receive":
			# stored = paid_to (company) → paid_from (foreign) = company → foreign.
			# source = paid_from (foreign) → company = 1 / stored
			source_exchange_rate = flt(1.0 / rate, 9) if rate else 1.0
		else:
			# Pay: paid_to = party = company; stored = paid_from (foreign) → paid_to (company).
			# source = paid_from → company = stored ✓
			source_exchange_rate = rate
	else:
		# Edge case: both accounts in non-company currencies.
		source_exchange_rate = rate
		target_exchange_rate_pe = rate

	# Derive each side's amount via the shared base (company-currency) amount.
	if payment_type == "Receive":
		# row_paid_amount is in paid_to_currency (cheque/bank currency entered by user).
		received_amount = raw_amount
		base_amount = received_amount * target_exchange_rate_pe
		paid_amount = flt(
			base_amount / source_exchange_rate if source_exchange_rate else base_amount, 9
		)
	else:
		# Pay: row_paid_amount is in paid_from_currency (cheque/bank currency).
		paid_amount = raw_amount
		base_amount = paid_amount * source_exchange_rate
		received_amount = flt(
			base_amount / target_exchange_rate_pe if target_exchange_rate_pe else base_amount, 9
		)

	return dict(
		paid_from_account_currency=paid_from_currency,
		paid_to_account_currency=paid_to_currency,
		paid_amount=paid_amount,
		received_amount=received_amount,
		source_exchange_rate=source_exchange_rate,
		target_exchange_rate=target_exchange_rate_pe,
	)


@frappe.whitelist()
def create_payment_entry_from_cheque(docname, row_id):
	"""Create and submit a single Payment Entry from a Cheque Table row.

	Fetches authoritative data from the database using *docname* (parent
	Multiple Cheque Entry) and *row_id* (child row ``name``).  Account
	currencies are read from the Account master so that stale UI values
	cannot cause incorrect exchange-rate assignments.

	ERPNext v15 multi-currency conventions for **Receive** (ILS company):
	  paid_amount          = row.amount_in_company_currency  (ILS)
	  received_amount      = row.paid_amount                 (USD)
	  source_exchange_rate = 1.0   (paid_from == company currency)
	  target_exchange_rate = row.target_exchange_rate        (USD→ILS)

	Returns the name of the submitted Payment Entry.
	"""
	doc = frappe.get_doc("Multiple Cheque Entry", docname)

	company = doc.company
	payment_type = doc.payment_type
	is_receive = payment_type == "Receive"

	company_currency = (
		frappe.db.get_value("Company", company, "default_currency") or ""
	)

	# Locate the child row in the appropriate table.
	child_doctype = "Cheque Table Receive" if is_receive else "Cheque Table Pay"
	table = doc.cheque_table if is_receive else doc.cheque_table_2
	row = next((r for r in (table or []) if r.name == row_id), None)
	if not row:
		frappe.throw(
			_("Child row {0} not found in document {1}").format(row_id, docname)
		)

	paid_from = row.account_paid_from
	paid_to = row.account_paid_to

	# Always fetch currencies from the Account master.
	paid_from_currency = _get_account_currency_db(paid_from, company_currency)
	paid_to_currency = _get_account_currency_db(paid_to, company_currency)

	raw_exchange_rate = flt(row.target_exchange_rate)
	if paid_from_currency != paid_to_currency and raw_exchange_rate <= 0:
		frappe.throw(
			_(
				"Row {0}: Cannot create Payment Entry — Exchange Rate is missing or zero "
				"for {1} → {2}. Please add a Currency Exchange record and retry."
			).format(
				row.idx or "",
				paid_to_currency if is_receive else paid_from_currency,
				paid_from_currency if is_receive else paid_to_currency,
			)
		)

	stored_rate = raw_exchange_rate or 1.0

	# Derive amounts and exchange rates directly from DB-stored values.
	if paid_from_currency == paid_to_currency:
		if paid_from_currency == company_currency:
			# Both accounts in company currency – no conversion needed.
			source_exchange_rate = 1.0
			target_exchange_rate = 1.0
			# Clear any stale exchange_rate_party_to_mop so that
			# _get_cheque_paid_amount won't trigger a false mismatch error.
			if is_receive and getattr(row, "exchange_rate_party_to_mop", None):
				frappe.db.set_value("Cheque Table Receive", row.name,
									"exchange_rate_party_to_mop", 0)
			# Bug 2: when the cheque is in a foreign currency (e.g. JOD) but
			# both accounts are in company currency (e.g. USD), paid_amount and
			# received_amount must be the company-currency equivalent so that
			# ERPNext does not create a spurious Exchange Gain/Loss entry.
			cheque_currency = getattr(row, "cheque_currency", None) or ""
			if cheque_currency and cheque_currency != company_currency:
				rate = flt(row.target_exchange_rate) or 1.0
				paid_amount = flt(row.paid_amount) * rate
			else:
				paid_amount = flt(row.paid_amount)
			received_amount = paid_amount
		else:
			# Both accounts share the same non-company currency (e.g. both USD).
			# Do not force rates to 1; leave them unset so ERPNext's validate()
			# will fetch the correct foreign-currency → company-currency rate.
			source_exchange_rate = None
			target_exchange_rate = None
			paid_amount = flt(row.paid_amount)
			received_amount = flt(row.paid_amount)
			# Clear any misleading exchange_rate_party_to_mop (e.g. 1.0 set by JS)
			# so that _get_cheque_paid_amount won't use the bidirectional rate path.
			if is_receive and getattr(row, "exchange_rate_party_to_mop", None):
				frappe.db.set_value("Cheque Table Receive", row.name,
									"exchange_rate_party_to_mop", 0)
	elif is_receive:
		# paid_from = party account, paid_to = bank/MOP account.
		paid_amount = flt(row.amount_in_company_currency)   # in paid_from currency
		received_amount = flt(row.paid_amount)              # in paid_to currency

		if paid_from_currency != company_currency and paid_to_currency != company_currency:
			# Both accounts are in non-company currencies (e.g. party=ILS, bank=JOD,
			# company=USD).
			#
			# Strict mapping: paid_amount (party account currency, e.g. ILS) must be
			# taken directly from amount_in_company_currency stored on the cheque row
			# rather than re-derived from exchange rates.  This prevents ERPNext from
			# replacing the fixed value (e.g. 3,000 ILS) with a system-recalculated
			# figure (e.g. 4,455 ILS).
			#
			# Only the MOP→company rate (target) is fetched from Currency Exchange.
			# The party→company rate (source) is derived so the GL equation balances:
			#   paid_amount × source_rate = received_amount × target_rate = base (USD)
			target_exchange_rate = _fetch_exchange_rate_to_company(
				paid_to_currency, company_currency, doc.posting_date
			)
			if not target_exchange_rate:
				frappe.throw(
					_(
						"Row {0}: No Currency Exchange record found for {1} → {2}. "
						"Please create one before proceeding."
					).format(row.idx or "", paid_to_currency, company_currency)
				)
			# paid_amount is the fixed party-account-currency amount from the cheque tool.
			paid_amount = flt(row.amount_in_company_currency)
			# base_company = received_amount (JOD) × target_rate (JOD → USD)
			base_company = flt(received_amount) * flt(target_exchange_rate)
			# Derive source_exchange_rate so that paid_amount × source_rate = base_company.
			source_exchange_rate = flt(base_company / paid_amount, 9) if paid_amount else 1.0
		elif paid_from_currency == company_currency:
			# paid_from is company currency → source rate must be 1.
			# target_exchange_rate converts paid_to (foreign) → company_currency.
			source_exchange_rate = 1.0
			target_exchange_rate = flt(paid_amount / received_amount, 9) if received_amount > 0 else stored_rate
		else:
			# paid_to_currency == company_currency
			source_exchange_rate = flt(1.0 / stored_rate, 9) if stored_rate else 1.0
			target_exchange_rate = 1.0
	else:
		# Pay: paid_from = bank/MOP account, paid_to = party account.
		paid_amount = flt(row.paid_amount)                     # in paid_from currency
		received_amount = flt(row.amount_in_company_currency)  # in paid_to currency

		if paid_from_currency != company_currency and paid_to_currency != company_currency:
			# Both accounts are in non-company currencies (e.g. bank=JOD, party=ILS,
			# company=USD).
			#
			# Strict mapping: received_amount (party account currency, e.g. ILS) must be
			# taken directly from amount_in_company_currency stored on the cheque row
			# rather than re-derived from exchange rates.
			#
			# Only the MOP→company rate (source) is fetched from Currency Exchange.
			# The party→company rate (target) is derived so the GL equation balances:
			#   paid_amount × source_rate = received_amount × target_rate = base (USD)
			source_exchange_rate = _fetch_exchange_rate_to_company(
				paid_from_currency, company_currency, doc.posting_date
			)
			if not source_exchange_rate:
				frappe.throw(
					_(
						"Row {0}: No Currency Exchange record found for {1} → {2}. "
						"Please create one before proceeding."
					).format(row.idx or "", paid_from_currency, company_currency)
				)
			# received_amount is the fixed party-account-currency amount from the cheque tool.
			received_amount = flt(row.amount_in_company_currency)
			# base_company = paid_amount (JOD) × source_rate (JOD → USD)
			base_company = flt(paid_amount) * flt(source_exchange_rate)
			# Derive target_exchange_rate so that received_amount × target_rate = base_company.
			target_exchange_rate = flt(base_company / received_amount, 9) if received_amount else 1.0
		elif paid_from_currency == company_currency:
			# Bank account is in company currency; party account is foreign.
			source_exchange_rate = 1.0
			target_exchange_rate = flt(received_amount / paid_amount, 9) if paid_amount > 0 else stored_rate
		else:
			# paid_to_currency == company_currency (party account = company currency)
			source_exchange_rate = stored_rate   # paid_from → company
			target_exchange_rate = 1.0

	pe_dict = {
		"doctype": "Payment Entry",
		"posting_date": doc.posting_date,
		"reference_doctype": "Multiple Cheque Entry",
		"reference_link": docname,
		"payment_type": payment_type,
		"company": company,
		"mode_of_payment": row.mode_of_payment or doc.mode_of_payment,
		"mode_of_payment_type": doc.mode_of_payment_type,
		"party_type": row.party_type,
		"party": row.party,
		"paid_from": paid_from,
		"paid_to": paid_to,
		"paid_from_account_currency": paid_from_currency,
		"paid_to_account_currency": paid_to_currency,
		"paid_amount": paid_amount,
		"received_amount": received_amount,
		"cheque_bank": doc.cheque_bank,
		"bank_acc": doc.bank_acc,
		"cheque_type": row.cheque_type,
		"reference_no": row.reference_no,
		"reference_date": row.reference_date,
		"first_beneficiary": row.first_beneficiary,
		"person_name": row.person_name,
		"issuer_name": row.issuer_name,
		"picture_of_check": row.picture_of_check,
		"cheque_table_no": row.name if is_receive else None,
		"cheque_table_no2": row.name if not is_receive else None,
	}

	# Only set exchange rates when they have a concrete value; leaving them
	# unset (None) lets ERPNext's validate() fetch the correct rate for
	# non-company-currency accounts.
	if source_exchange_rate is not None:
		pe_dict["source_exchange_rate"] = source_exchange_rate
	if target_exchange_rate is not None:
		pe_dict["target_exchange_rate"] = target_exchange_rate

	if is_receive:
		pe_dict["drawn_bank"] = row.bank

	pe_doc = frappe.get_doc(pe_dict)
	pe_doc.flags.ignore_permissions = True
	pe_doc.insert()
	pe_doc.submit()

	# Persist the Payment Entry link back to the child row in the database.
	frappe.db.set_value(child_doctype, row_id, "payment_entry", pe_doc.name)

	return pe_doc.name


class MultipleChequeEntry(Document):
	def before_save(self):
		"""Compute amount_in_usd for all child rows before saving."""
		self._compute_amount_in_usd_all_rows()

	def _compute_amount_in_usd_all_rows(self):
		"""Calculate amount_in_usd (amount in company currency) for every child row.

		Fetches the company currency once, then for each cheque row converts
		``paid_amount`` from ``cheque_currency`` → company currency using the
		Currency Exchange DocType.  Results are stored directly on the row
		objects so they are saved with the document.
		"""
		company_currency = (
			frappe.db.get_value("Company", self.company, "default_currency") or ""
			if self.company else ""
		)
		if not company_currency:
			return

		posting_date = self.posting_date or nowdate()
		tables = []
		if self.payment_type == "Receive":
			tables.append((self.cheque_table or [], "Cheque Table Receive", "account_currency"))
		else:
			tables.append((self.cheque_table_2 or [], "Cheque Table Pay", "account_currency_from"))

		for rows, _child_doctype, currency_field in tables:
			for row in rows:
				cheque_currency = getattr(row, "cheque_currency", None) or getattr(row, currency_field, None) or ""
				paid_amount = flt(getattr(row, "paid_amount", 0))
				if not cheque_currency or cheque_currency == company_currency:
					row.amount_in_usd = paid_amount
				else:
					rate = _fetch_exchange_rate_to_company(cheque_currency, company_currency, posting_date)
					row.amount_in_usd = flt(paid_amount * (rate or 1.0), 9)

	def on_cancel(self):
		"""Cancel linked Payment Entries when Multiple Cheque Entry is cancelled."""
		pe_names = self._get_linked_payment_entries()
		for pe_name in pe_names:
			try:
				pe = frappe.get_doc("Payment Entry", pe_name)
				if pe.docstatus == 1:
					pe.cancel()
			except Exception as e:
				frappe.throw(_("Failed to cancel Payment Entry {0}: {1}").format(pe_name, str(e)))

	def on_trash(self):
		"""Delete linked Payment Entries when Multiple Cheque Entry is deleted."""
		pe_names = self._get_linked_payment_entries()
		for pe_name in pe_names:
			try:
				pe = frappe.get_doc("Payment Entry", pe_name)
				if pe.docstatus == 2:
					frappe.delete_doc("Payment Entry", pe_name, ignore_permissions=True)
			except Exception as e:
				frappe.throw(_("Failed to delete Payment Entry {0}: {1}").format(pe_name, str(e)))

	def _get_linked_payment_entries(self):
		"""Return list of Payment Entry names linked to this document."""
		return frappe.get_all(
			"Payment Entry",
			filters={"reference_doctype": "Multiple Cheque Entry", "reference_link": self.name},
			pluck="name"
		)


@frappe.whitelist()
def get_cheques_excel_template(payment_type):
	"""Return an Excel (.xlsx) template with headers and 3 sample rows for the cheque table."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw(_("openpyxl is required to generate Excel templates. Please install it."))

	wb = openpyxl.Workbook()
	ws = wb.active

	if payment_type == "Receive":
		ws.title = "Cheques Receive"
		headers = [
			"party_type", "party", "mode_of_payment", "bank",
			"reference_no", "reference_date", "cheque_type",
			"cheque_currency", "paid_amount", "target_exchange_rate",
			"account_paid_from", "account_paid_to",
			"first_beneficiary", "person_name", "issuer_name"
		]
		samples = [
			["Customer", "CUST-001", "شيك", "National Bank",
			 "CHQ-001", "2024-01-15", "Crossed",
			 "EGP", 5000, 1,
			 "1310 - مدينون - O", "1110 - بنك - O",
			 "Company", "", "Ahmed Ali"],
			["Customer", "CUST-002", "شيك", "ABC Bank",
			 "CHQ-002", "2024-01-20", "Opened",
			 "USD", 1000, 48.5,
			 "1310 - مدينون - O", "1111 - بنك دولار - O",
			 "Personal", "Mohamed Said", "Mohamed Said"],
			["Supplier", "SUPP-001", "شيك", "National Bank",
			 "CHQ-003", "2024-02-01", "Crossed",
			 "EGP", 12000, 1,
			 "2110 - دائنون - O", "1110 - بنك - O",
			 "", "", ""],
		]
	else:
		ws.title = "Cheques Pay"
		headers = [
			"party_type", "party", "mode_of_payment",
			"reference_no", "reference_date", "cheque_type",
			"cheque_currency", "paid_amount", "target_exchange_rate",
			"account_paid_from", "account_paid_to",
			"first_beneficiary", "person_name", "issuer_name"
		]
		samples = [
			["Supplier", "SUPP-001", "شيك",
			 "CHQ-PAY-001", "2024-01-15", "Crossed",
			 "EGP", 8000, 1,
			 "1110 - بنك - O", "2110 - دائنون - O",
			 "", "", ""],
			["Supplier", "SUPP-002", "شيك",
			 "CHQ-PAY-002", "2024-01-22", "Opened",
			 "USD", 500, 48.5,
			 "1111 - بنك دولار - O", "2110 - دائنون - O",
			 "", "", ""],
			["Customer", "CUST-001", "شيك",
			 "CHQ-PAY-003", "2024-02-05", "Crossed",
			 "EGP", 3000, 1,
			 "1110 - بنك - O", "1310 - مدينون - O",
			 "", "", ""],
		]

	ws.append(headers)
	for sample in samples:
		ws.append(sample)

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response.filename = "cheques_template_{}.xlsx".format(payment_type.lower())
	frappe.response.filecontent = output.read()
	frappe.response.type = "binary"


@frappe.whitelist()
def upload_cheques_excel(file_data, payment_type):
	"""Parse uploaded Excel file and return rows as list of dicts for the cheque table."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw(_("openpyxl is required to upload Excel files. Please install it."))

	if isinstance(file_data, str):
		import base64
		file_bytes = base64.b64decode(file_data)
	else:
		file_bytes = file_data

	wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
	ws = wb.active

	rows = list(ws.iter_rows(values_only=True))
	if not rows:
		frappe.throw(_("Excel file is empty."))

	headers = [str(h).strip() if h else "" for h in rows[0]]
	required_cols = ["party_type", "party", "reference_no", "reference_date",
					 "cheque_type", "paid_amount"]

	missing = [c for c in required_cols if c not in headers]
	if missing:
		frappe.throw(_("Missing required columns: {0}").format(", ".join(missing)))

	result = []
	errors = []
	for i, row in enumerate(rows[1:], start=2):
		row_dict = {headers[j]: row[j] for j in range(len(headers)) if j < len(row)}

		# Basic validation
		for col in required_cols:
			if not row_dict.get(col):
				errors.append(_("Row {0}: '{1}' is required.").format(i, col))

		paid_amount = row_dict.get("paid_amount")
		if paid_amount is not None:
			try:
				row_dict["paid_amount"] = flt(paid_amount)
				if row_dict["paid_amount"] <= 0:
					errors.append(_("Row {0}: 'paid_amount' must be greater than zero.").format(i))
			except Exception:
				errors.append(_("Row {0}: 'paid_amount' must be a number.").format(i))

		exchange_rate = row_dict.get("target_exchange_rate")
		if exchange_rate is not None:
			try:
				row_dict["target_exchange_rate"] = flt(exchange_rate) or 1
			except Exception:
				row_dict["target_exchange_rate"] = 1

		# Convert date to string
		ref_date = row_dict.get("reference_date")
		if ref_date and not isinstance(ref_date, str):
			row_dict["reference_date"] = str(ref_date.date()) if hasattr(ref_date, 'date') else str(ref_date)

		result.append(row_dict)

	if errors:
		frappe.throw("<br>".join(errors))

	return result
